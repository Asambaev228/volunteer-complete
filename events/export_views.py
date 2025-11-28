import openpyxl
from django.http import HttpResponse
from django.contrib.admin.views.decorators import staff_member_required
from .models import Event, EventCategory, VolunteerApplication

@staff_member_required
def export_events_excel(request):
    """Экспорт мероприятий в Excel"""
    events = Event.objects.all().select_related('category', 'organizer')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Мероприятия"
    
    headers = ['ID', 'Название', 'Категория', 'Организатор', 'Дата', 'Место', 'Статус', 'Нужно волонтеров', 'Создано']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    for row, event in enumerate(events, 2):
        ws.cell(row=row, column=1, value=event.id)
        ws.cell(row=row, column=2, value=event.title)
        ws.cell(row=row, column=3, value=str(event.category))
        ws.cell(row=row, column=4, value=str(event.organizer))
        ws.cell(row=row, column=5, value=event.event_date.strftime('%d.%m.%Y %H:%M'))
        ws.cell(row=row, column=6, value=event.location)
        ws.cell(row=row, column=7, value=event.get_status_display())
        ws.cell(row=row, column=8, value=event.required_volunteers)
        ws.cell(row=row, column=9, value=event.created_at.strftime('%d.%m.%Y %H:%M'))
    
    # Настраиваем ширину
    for column in ws.columns:
        max_length = 0
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[openpyxl.utils.get_column_letter(column[0].column)].width = adjusted_width
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=events_export.xlsx'
    wb.save(response)
    return response

@staff_member_required
def export_categories_excel(request):
    """Экспорт категорий в Excel"""
    categories = EventCategory.objects.all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Категории"
    
    headers = ['ID', 'Название', 'Описание', 'Создано', 'Обновлено']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    for row, category in enumerate(categories, 2):
        ws.cell(row=row, column=1, value=category.id)
        ws.cell(row=row, column=2, value=category.name)
        ws.cell(row=row, column=3, value=category.description)
        ws.cell(row=row, column=4, value=category.created_at.strftime('%d.%m.%Y %H:%M'))
        ws.cell(row=row, column=5, value=category.updated_at.strftime('%d.%m.%Y %H:%M'))
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=categories_export.xlsx'
    wb.save(response)
    return response

@staff_member_required
def export_applications_excel(request):
    """Экспорт заявок в Excel"""
    applications = VolunteerApplication.objects.all().select_related('volunteer', 'event')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Заявки"
    
    headers = ['ID', 'Волонтер', 'Мероприятие', 'Статус', 'Сообщение', 'Создано']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    for row, application in enumerate(applications, 2):
        ws.cell(row=row, column=1, value=application.id)
        ws.cell(row=row, column=2, value=str(application.volunteer))
        ws.cell(row=row, column=3, value=str(application.event))
        ws.cell(row=row, column=4, value=application.get_status_display())
        ws.cell(row=row, column=5, value=application.message)
        ws.cell(row=row, column=6, value=application.created_at.strftime('%d.%m.%Y %H:%M'))
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=applications_export.xlsx'
    wb.save(response)
    return response