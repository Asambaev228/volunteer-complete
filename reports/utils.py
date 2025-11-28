import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from events.models import Event, EventCategory, VolunteerApplication
from django.contrib.auth.models import User

def export_to_excel(model_name, fields):
    """
    Создает Excel файл с данными указанной модели.
    
    Args:
        model_name (str): Название модели для экспорта
        fields (list): Список полей для включения в отчет
    
    Returns:
        openpyxl.Workbook: Готовый Excel файл
    """
    # Создаем новую книгу Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Определяем данные для экспорта
    if model_name == 'events':
        data = Event.objects.all()
        ws.title = "Мероприятия"
    elif model_name == 'categories':
        data = EventCategory.objects.all()
        ws.title = "Категории мероприятий"
    elif model_name == 'applications':
        data = VolunteerApplication.objects.all()
        ws.title = "Заявки волонтеров"
    elif model_name == 'users':
        data = User.objects.all()
        ws.title = "Пользователи"
    else:
        return None
    
    # Добавляем заголовки столбцов
    for col_num, field_name in enumerate(fields, 1):
        col_letter = get_column_letter(col_num)
        # Преобразуем названия полей в читаемый формат
        readable_name = field_name.replace('_', ' ').title()
        ws[f'{col_letter}1'] = readable_name
    
    # Заполняем данными
    for row_num, obj in enumerate(data, 2):
        for col_num, field_name in enumerate(fields, 1):
            col_letter = get_column_letter(col_num)
            value = getattr(obj, field_name)
            
            # Обрабатываем специальные типы полей
            if hasattr(value, 'username'):  # Пользователь
                value = value.username
            elif hasattr(value, 'name'):    # Категория
                value = value.name
            elif hasattr(value, 'title'):   # Мероприятие
                value = value.title
            elif hasattr(value, 'get_status_display'):  # Choice поле
                value = getattr(obj, f'get_{field_name}_display')()
            
            ws[f'{col_letter}{row_num}'] = str(value) if value else ''
    
    # Настраиваем ширину столбцов
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Максимальная ширина 50
        ws.column_dimensions[column_letter].width = adjusted_width
    
    return wb

def export_selected_objects(queryset, model_name):
    """
    Экспорт выбранных объектов в Excel.
    """
    # Определяем поля для экспорта в зависимости от модели
    field_mapping = {
        'events': ['title', 'category', 'organizer', 'event_date', 'location', 'status', 'created_at'],
        'categories': ['name', 'description', 'created_at', 'updated_at'],
        'applications': ['volunteer', 'event', 'status', 'message', 'created_at'],
        'users': ['username', 'email', 'first_name', 'last_name', 'is_staff', 'date_joined']
    }
    
    fields = field_mapping.get(model_name, [])
    return export_to_excel(model_name, fields)